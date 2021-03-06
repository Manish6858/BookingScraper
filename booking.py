#! /usr/bin/env python3

# Uncomment for automatic installation of the requirements
#import sys
#sys.path.insert(0, "./lib/python3.5/site-packages")
#import os
#os.system('pip install -r ./requirements.txt')

import requests
from bs4 import BeautifulSoup


def get_booking_page(offset):
    '''
    Make request to booking page and parse html
    :param offset:
    :return: html page
    '''
    url = 'https://www.booking.com/searchresults.en-gb.html?' \
            'aid=304142&label' \
            '=gen173nr-1DCAEoggJCAlhYSDNYBGiTAYgBAZgBLsIBCnd' \
            'pbmRvd3MgMTDIAQzYAQPoAQGSAgF5qAID&' \
            'sid=716ea5d78c4043fd78b7a1410d639e3d&checkin_month=' \
            '6&checkin_monthday=8&checkin_year=2018&checkout_month=6&' \
            'checkout_monthday=11&checkout_year=2018' \
            '&class_interval=1&dest_id=125&dest_type=country&dtdisc=0&from_sf'\
            '=1&genius_rate=1&group_adults=2&group_children=0&inac=0&' \
            'index_postcard=0&label_click=undef' \
            '&no_rooms=1&postcard=0&raw_dest_type=country&room1=' \
            'A%2CA&sb_price' \
            '_type=total&src=searchresults&src_elem=sb&ss=Macedonia&ss_all=' \
            '0&ssb=empty&sshis=0&ssne=Macedonia' \
            '&ssne_untouched=Macedonia&rows=15&offset=' + str(offset)
    r = requests.get(url, headers=
      {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:47.0)'
                     ' Gecko/20100101 Firefox/48.0'})
    html = r.content
    parsed_html = BeautifulSoup(html, 'lxml')
    return parsed_html


def get_data():
    '''
    Get all accomodations in Macedonia and save them in file
    :return: hotels-in-macedonia.{txt/csv/xlsx} file
    '''
    offset = 15
    parsed_html = get_booking_page(offset)
    all_offset = parsed_html.find_all('li', {'class':
                                      'sr_pagination_item'})[-1].get_text()

    hotels = set()
    number = 0
    for i in range(int(all_offset)):
        offset += 15
        number+=1
        parsed_html = get_booking_page(offset)
        hotel = parsed_html.find_all('div', {'class': 'sr_item'})

        for ho in hotel:
            name = ho.find('a', {'class': 'jq_tooltip'})['title']
            hotels.add(str(number) + ' : ' + name)
            number += 1

    save_data(hotels)


def save_data(data, out_format=None):
    '''
    Saves hotels list in file
    :param data: hotels list
    :param out_format: json, csv or excel
    :return:
    '''
    if out_format is None:
        import json
        file_name = 'hotels-in-macedonia.txt'
        with open(file_name, 'w', encoding='utf-8') as outfile:
            json.dump(list(data), outfile, indent=2, ensure_ascii=False)

    elif out_format == 'excel':
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active

        heading1 = '#'
        heading2 = 'Accommodation'
        ws.cell(row=1, column=1).value = heading1
        ws.cell(row=1, column=2).value = heading2

        for i, item in enumerate(data):
            # Extract number and title from string
            tokens = item.split()
            n = tokens[0]
            title = ' '.join(tokens[2:])

            ws.cell(row=i + 2, column=1).value = n
            ws.cell(row=i + 2, column=2).value = title

        file_name = 'hotels-in-macedonia.xlsx'
        wb.save(file_name)

    elif out_format == 'csv':
        file_name = 'hotels-in-macedonia.csv'
        with open(file_name, 'w', encoding='utf-8') as outfile:
            for i, item in enumerate(data):
                # Extract number and title from string
                tokens = item.split()
                n = tokens[0]
                title = ' '.join(tokens[2:])

                s = n + ', ' + title + '\n'
                outfile.write(s)


    print('All accommodations are saved.')
    print('You can find them in', file_name, 'file')

if __name__ == "__main__":
    get_data()
